import sys
sys.path.append("./")
from Library.GlobalAdapter import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import unittest
from Core.ApiQuery import*
from Library.DBConnector import*
from Utility.testrail import*
import configparser
class Newconfigparser(configparser.ConfigParser):
    def __init__(self,defaults=None):
        configparser.ConfigParser.__init__(self,defaults=None)
    def optionxform(self, optionstr):
        return optionstr

class Order_flow(unittest.TestCase):
    @classmethod
    def setUpClass(self):
        self.TestSuiteName = "Service Offering"
        self.runId, self.testIds = AddTestRunAndTest(self.TestSuiteName)
        self.TestCaseName = testrail_config[self.TestSuiteName]['case_names'].split(',')
        self.wb = load_workbook('./Config/ServiceOfferingCondition.xlsx')
        self.sheet = self.wb.active

    def tearDown(self):
        dumplogger.info("Recovery : Start to Delete Data for this TestCase")
        #Delete Service Offering
        for ServiceOffering_id in self.DSO:
            DeleteServiceOffering(ServiceOffering_id)
        #Delete Timeslot
        for Timeslot_id in self.DTS:
            DeleteTimeslot(Timeslot_id)
        #Delete OrderFlow
        for OrderFlow_id in self.DOF:
            DeleteOrderFlow(OrderFlow_id)
        #Delete Price Model
        for Price_id in self.DPM:
            DeletePriceModel(Price_id)

    @classmethod
    def tearDownClass(self):
        GetRunResultAndCheckStatusThenSendToSlack(self.runId,self.TestSuiteName)

    def CheckTotalStatus(self,resultList):
        for index in resultList:
            if index == False:
                return 5
        return 1   

    def TranslateForStatus(self,resultList):
        FinalResult = []
        for index in resultList:
            if index == True:
                FinalResult.append(1)
            elif index == False:
                FinalResult.append(5)
            else:
                FinalResult.append(index)
        return FinalResult
    
    def test_01(self):
        maxcolumn = self.sheet.max_column
        self.DPM, self.DOF, self.DTS, self.DSO = [],[],[],[]
        for i in range(maxcolumn-1):
            totalstatus = True
            resultList = []
            columnletter = get_column_letter((i+2))
            #Price model
            PriceRange = self.sheet['%s2' %columnletter:'%s13' % columnletter]
            PriceConfig = []
            for price in PriceRange:
                for cell in price:
                    PriceConfig.append(cell.value)
            status_code, price_id, status = service_offering_price(PriceConfig,totalstatus)
            resultList.extend((status, price_id))
            self.DPM.append(price_id)
            #Order Flow
            OrderFlowRange = self.sheet['%s15' % columnletter].value
            status_code, orderflow_id, status = service_offering_orderflow(OrderFlowRange,totalstatus)
            resultList.extend((status, orderflow_id))
            self.DOF.append(orderflow_id)

            #Timeslot Early Pickup
            TimeslotEPRange = self.sheet['%s17' %columnletter:'%s22' % columnletter]
            TimeslotEPConfig = []
            for timeslot in TimeslotEPRange:
                for cell in timeslot:
                    TimeslotEPConfig.append(cell.value)
            status_code, timeslot_ep, status = service_offering_timeslot(TimeslotEPConfig,totalstatus)
            resultList.extend((status, timeslot_ep))
            self.DTS.append(timeslot_ep)

            #Timeslot Last Pickup
            TimeslotLPRange = self.sheet['%s24' %columnletter:'%s29' % columnletter]
            TimeslotLPConfig = []
            for timeslot in TimeslotLPRange:
                for cell in timeslot:
                    TimeslotLPConfig.append(cell.value)
            TimeslotLPConfig[2] = timeslot_ep
            status_code, timeslot_lp, status = service_offering_timeslot(TimeslotLPConfig,totalstatus)
            resultList.extend((status, timeslot_lp))
            self.DTS.append(timeslot_lp)

            #Timeslot Early Dropoff           
            TimeslotEDRange = self.sheet['%s31' %columnletter:'%s36' % columnletter]
            TimeslotEDConfig = []
            for timeslot in TimeslotEDRange:
                for cell in timeslot:
                    TimeslotEDConfig.append(cell.value)
            TimeslotEDConfig[2] = timeslot_ep
            status_code, timeslot_ed, status = service_offering_timeslot(TimeslotEDConfig,totalstatus)
            resultList.extend((status, timeslot_ed))
            self.DTS.append(timeslot_ed)

            #Timeslot Early Last Dropoff 
            TimeslotLDRange = self.sheet['%s38' %columnletter:'%s43' % columnletter]
            TimeslotLDConfig = []
            for timeslot in TimeslotLDRange:
                for cell in timeslot:
                    TimeslotLDConfig.append(cell.value)
            TimeslotLDConfig[2] = timeslot_lp
            status_code, timeslot_ld, status = service_offering_timeslot(TimeslotLDConfig,totalstatus)
            resultList.extend((status, timeslot_ld))
            self.DTS.append(timeslot_ld)

            #Timeslot Percentage
            TimeslotPerRange = self.sheet['%s45' %columnletter:'%s50' % columnletter]
            TimeslotPerConfig = []
            for timeslot in TimeslotPerRange:
                for cell in timeslot:
                    TimeslotPerConfig.append(cell.value)
            TimeslotPerConfig[2:3] = [timeslot_lp,timeslot_ld]
            status_code, timeslot_per, status = service_offering_timeslot(TimeslotPerConfig,totalstatus)
            resultList.extend((status, timeslot_per))
            self.DTS.append(timeslot_per)

            #Service Offering
            ServiceOfferingRange = self.sheet['%s52' %columnletter:'%s67' % columnletter]
            ServiceOfferingConfig = []
            for timeslot in ServiceOfferingRange:
                for cell in timeslot:
                    ServiceOfferingConfig.append(cell.value)
            ServiceOfferingConfig[5:10] = [timeslot_ep,timeslot_lp,timeslot_ed,timeslot_ld,price_id,orderflow_id]
            status_code, serviceoffering_id, status = service_offering_service(ServiceOfferingConfig,totalstatus)
            resultList.extend((status, serviceoffering_id))
            self.DSO.append(serviceoffering_id)

            #Update result to Testrail
            totalstatus = self.CheckTotalStatus(resultList)
            resultList = self.TranslateForStatus(resultList)
            dumplogger.info(self.TestCaseName[i])
            AddResultByStep(totalstatus,self.testIds,self.TestCaseName[i],resultList,self.runId)    

if __name__ == '__main__':
    unittest.main()  

